import tkinter as tk
from tkinter import simpledialog, messagebox, filedialog
from openpyxl import load_workbook
import datetime
import webbrowser
import uuid
import os

# Dictionary to convert an integer month to its Croatian name.
CRO_MONTHS = {
    1: "Siječanj",
    2: "Veljača",
    3: "Ožujak",
    4: "Travanj",
    5: "Svibanj",
    6: "Lipanj",
    7: "Srpanj",
    8: "Kolovoz",
    9: "Rujan",
    10: "Listopad",
    11: "Studeni",
    12: "Prosinac"
}

# Location used for timed events.
LOCATION = "Klinički bolnički centar Zagreb, Ulica Mije Kišpatića 12, 10000, Zagreb"

def main():
    # 1) Create a hidden root window for Tk dialogs
    root = tk.Tk()
    root.withdraw()

    # 2) Ask user to select the Excel file
    excel_file = filedialog.askopenfilename(
        title="Odaberi Excel datoteku",
        filetypes=[("Excel datoteke", "*.xlsx *.xls"), ("Sve datoteke", "*.*")]
    )

    if not excel_file:
        print("Nije odabrana datoteka, pokušaj opet.")
        return

    # --- Identify the directory of the selected Excel file ---
    excel_dir = os.path.dirname(excel_file)

    # 3) Load the selected Excel workbook
    try:
        wb = load_workbook(excel_file)
    except Exception as e:
        print(f"Pogreška pri učitavanju Excel datoteke: {e}")
        return

    # If you have a specific worksheet name, use wb["SheetName"], otherwise use the active sheet.
    sheet = wb.active

    # 4) Prompt user input in a popup window for the surname
    valid_surname = False
    surname_row = None

    while not valid_surname:
        surname_input = simpledialog.askstring(
            "Unos prezimena",
            "Unesi svoje prezime (kako si inače naveden u rasporedu):"
        )

        if surname_input is None:
            # User hit cancel or closed the dialog
            print("Nije uneseno prezime, probaj opet.")
            return

        # Try to find the input in Column A
        for row in range(1, sheet.max_row + 1):
            cell_value = sheet.cell(row=row, column=1).value
            if cell_value and str(cell_value).strip() == surname_input.strip():
                surname_row = row
                break

        if surname_row is None:
            messagebox.showerror(
                "Neispravan unos",
                "Neispravan unos, probaj opet."
            )
        else:
            valid_surname = True

    # 5) Determine the month name based on the first date in row=1, column=2
    first_date_cell = sheet.cell(row=1, column=2).value

    # Parse date (string "D.M.YYYY" or a real date object)
    if isinstance(first_date_cell, str):
        try:
            day_str, month_str, year_str = first_date_cell.split(".")
            day = int(day_str.strip())
            month = int(month_str.strip())
            year = int(year_str.strip())
        except Exception:
            print("Ne mogu parsirati prvi datum (ćelija B2). Provjeri format npr. '1.2.2025'.")
            return
    elif isinstance(first_date_cell, datetime.date):
        day = first_date_cell.day
        month = first_date_cell.month
        year = first_date_cell.year
    else:
        print("Nepriznat format datuma u ćeliji B2.")
        return

    croatian_month_name = CRO_MONTHS.get(month, "Nepoznat_mjesec")

    # 6) Create ICS text
    ics_lines = []
    ics_lines.append("BEGIN:VCALENDAR")
    ics_lines.append("VERSION:2.0")
    ics_lines.append("PRODID:-//MyOrg//Schedule Export//EN")

    # We'll iterate through columns starting at column=2, where each column=day of the month
    max_col = sheet.max_column

    for col in range(2, max_col + 1):
        date_cell = sheet.cell(row=1, column=col).value
        if not date_cell:
            # No date => skip
            continue

        # Parse the date in row=1, col=col
        if isinstance(date_cell, str):
            # Expect format like "15.2.2025"
            try:
                d_str, m_str, y_str = date_cell.split(".")
                day_d = int(d_str.strip())
                month_d = int(m_str.strip())
                year_d = int(y_str.strip())
                date_obj = datetime.date(year_d, month_d, day_d)
            except Exception:
                # Skip if format is invalid
                continue
        elif isinstance(date_cell, datetime.date):
            date_obj = date_cell
        else:
            # Invalid or empty => skip
            continue

        schedule_value = sheet.cell(row=surname_row, column=col).value
        if not schedule_value:
            # Empty => no event
            continue

        schedule_str = str(schedule_value).strip()

        # Create a unique UID for each event so they never overwrite older imports
        uid_str = f"{surname_input}-{date_obj.year}-{date_obj.month}-{date_obj.day}-{col}-{uuid.uuid4()}@MyOrg"

        # If you want to fix accidental "G0" -> "GO", you could do:
        # if schedule_str.upper() == "G0":
        #     schedule_str = "GO"

        # Full-day events include "GO" (with letter O)
        if schedule_str in ["SD", "PD", "GO", "bol", "sist", "postdipl", "klaićeva"]:
            # All-day event
            dt_start = date_obj.strftime("%Y%m%d")
            dt_end_obj = date_obj + datetime.timedelta(days=1)
            dt_end = dt_end_obj.strftime("%Y%m%d")

            ics_lines.append("BEGIN:VEVENT")
            ics_lines.append(f"UID:{uid_str}")
            ics_lines.append(f"DTSTAMP:{datetime.datetime.now().strftime('%Y%m%dT%H%M%S')}")
            ics_lines.append(f"DTSTART;VALUE=DATE:{dt_start}")
            ics_lines.append(f"DTEND;VALUE=DATE:{dt_end}")
            ics_lines.append(f"SUMMARY:{schedule_str}")
            ics_lines.append("END:VEVENT")

        elif schedule_str.endswith("p"):
            # Ends with 'p' => 14:00 to 20:00
            start_dt = datetime.datetime(year=date_obj.year, month=date_obj.month, day=date_obj.day, hour=14, minute=0)
            end_dt = datetime.datetime(year=date_obj.year, month=date_obj.month, day=date_obj.day, hour=20, minute=0)

            dt_start_str = start_dt.strftime("%Y%m%dT%H%M%S")
            dt_end_str = end_dt.strftime("%Y%m%dT%H%M%S")

            ics_lines.append("BEGIN:VEVENT")
            ics_lines.append(f"UID:{uid_str}")
            ics_lines.append(f"DTSTAMP:{datetime.datetime.now().strftime('%Y%m%dT%H%M%S')}")
            ics_lines.append(f"DTSTART:{dt_start_str}")
            ics_lines.append(f"DTEND:{dt_end_str}")
            ics_lines.append(f"SUMMARY:{schedule_str}")
            ics_lines.append(f"LOCATION:{LOCATION}")
            ics_lines.append("END:VEVENT")

        elif schedule_str == "H16":
            # H16 => 15:30 to 07:30 next day
            start_dt = datetime.datetime(year=date_obj.year, month=date_obj.month, day=date_obj.day, hour=15, minute=30)
            next_day = date_obj + datetime.timedelta(days=1)
            end_dt = datetime.datetime(year=next_day.year, month=next_day.month, day=next_day.day, hour=7, minute=30)

            dt_start_str = start_dt.strftime("%Y%m%dT%H%M%S")
            dt_end_str = end_dt.strftime("%Y%m%dT%H%M%S")

            ics_lines.append("BEGIN:VEVENT")
            ics_lines.append(f"UID:{uid_str}")
            ics_lines.append(f"DTSTAMP:{datetime.datetime.now().strftime('%Y%m%dT%H%M%S')}")
            ics_lines.append(f"DTSTART:{dt_start_str}")
            ics_lines.append(f"DTEND:{dt_end_str}")
            ics_lines.append(f"SUMMARY:{schedule_str}")
            ics_lines.append(f"LOCATION:{LOCATION}")
            ics_lines.append("END:VEVENT")

        else:
            # Default => 07:30 to 15:00
            start_dt = datetime.datetime(year=date_obj.year, month=date_obj.month, day=date_obj.day, hour=7, minute=30)
            end_dt = datetime.datetime(year=date_obj.year, month=date_obj.month, day=date_obj.day, hour=15, minute=0)

            dt_start_str = start_dt.strftime("%Y%m%dT%H%M%S")
            dt_end_str = end_dt.strftime("%Y%m%dT%H%M%S")

            ics_lines.append("BEGIN:VEVENT")
            ics_lines.append(f"UID:{uid_str}")
            ics_lines.append(f"DTSTAMP:{datetime.datetime.now().strftime('%Y%m%dT%H%M%S')}")
            ics_lines.append(f"DTSTART:{dt_start_str}")
            ics_lines.append(f"DTEND:{dt_end_str}")
            ics_lines.append(f"SUMMARY:{schedule_str}")
            ics_lines.append(f"LOCATION:{LOCATION}")
            ics_lines.append("END:VEVENT")

    ics_lines.append("END:VCALENDAR")

    # 7) Write the ICS file in the same directory as the Excel file
    ics_filename = f"{surname_input} {croatian_month_name}.ics"
    ics_full_path = os.path.join(excel_dir, ics_filename)

    try:
        with open(ics_full_path, "w", encoding="utf-8") as f:
            for line in ics_lines:
                f.write(line + "\r\n")
        print(f"ICS datoteka '{ics_filename}' uspješno stvorena u direktoriju: {excel_dir}")
    except Exception as e:
        print(f"Pogreška pri pisanju ICS datoteke: {e}")
        return

    # 8) Open Google Calendar export page in Chrome
    url = "https://calendar.google.com/calendar/u/0/r/settings/export"

    # Attempt to open specifically in Chrome (Windows example path):
    chrome_path_64 = r"C:/Program Files/Google/Chrome/Application/chrome.exe %s"
    chrome_path_32 = r"C:/Program Files (x86)/Google/Chrome/Application/chrome.exe %s"

    try:
        # Try 64-bit path first
        webbrowser.get(chrome_path_64).open(url)
    except:
        try:
            # Fallback to 32-bit path
            webbrowser.get(chrome_path_32).open(url)
        except:
            # If that fails, open using the default browser
            webbrowser.open(url)
            print("Nije moguće pronaći Chrome na zadanom putu, otvaram u zadanom web-pregledniku...")

if __name__ == "__main__":
    main()
