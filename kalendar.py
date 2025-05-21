import tkinter as tk
from tkinter import simpledialog, messagebox, filedialog
from openpyxl import load_workbook
import datetime
import webbrowser
import uuid
import os

# Dictionary to convert an integer month to its Croatian name.
CRO_MONTHS = {
    1: "Siječanj", 2: "Veljača", 3: "Ožujak", 4: "Travanj", 5: "Svibanj", 6: "Lipanj",
    7: "Srpanj", 8: "Kolovoz", 9: "Rujan", 10: "Listopad", 11: "Studeni", 12: "Prosinac"
}

# Location used for timed events.
LOCATION = "Klinički bolnički centar Zagreb, Ulica Mije Kišpatića 12, 10000, Zagreb"

# Events to exclude from colleague matching
EXCLUDED_EVENTS_FOR_MATCHING = ["SD", "PD", "GO", "bol", "sist", "postdipl", "klaićeva"]


def is_weekend(date_obj):
    """Checks if the date is a Saturday (5) or Sunday (6)."""
    return date_obj.weekday() >= 5


def get_event_parts(event_str):
    """Splits an event string like 'A/B' into a set {'A', 'B'}, or 'A' into {'A'}."""
    if not event_str:
        return set()
    return set(s.strip() for s in str(event_str).strip().split('/'))


def is_match(user_event_raw, colleague_event_raw, date_obj):
    """
    Determines if the colleague's event matches the user's event based on specified rules.
    """
    user_event_str = str(user_event_raw).strip()
    colleague_event_str = str(colleague_event_raw).strip()

    if not user_event_str or not colleague_event_str:
        return False

    user_parts = get_event_parts(user_event_str)
    colleague_parts = get_event_parts(colleague_event_str)

    # --- Rule 1: H16 related matching (any day) ---
    user_is_h16 = (user_event_str == "H16")
    colleague_is_h16 = (colleague_event_str == "H16")
    user_has_dez_component = "dež" in user_parts
    colleague_has_dez_component = "dež" in colleague_parts

    if user_is_h16 and colleague_is_h16:  # H16 matches H16
        return True
    if (user_is_h16 and colleague_has_dez_component) or \
            (user_has_dez_component and colleague_is_h16):  # H16 <-> dež component
        return True

    # --- Rule 2: General H8 related matching (any day) ---
    # An event is "H8-type" if "H8" is one of its components.
    user_is_h8_type = "H8" in user_parts
    colleague_is_h8_type = "H8" in colleague_parts

    if user_is_h8_type and colleague_is_h8_type:
        # This means "H8" matches "H8".
        # "H8" matches "H8/dež", "H8/J" etc. (and vice-versa).
        # "H8/dež" matches "H8/J".
        return True

    # --- Rule 3: Additional Weekend "dež" matching (with H8-type or other dež-type) ---
    if is_weekend(date_obj):
        user_is_weekend_h8_or_dez_type = user_is_h8_type or user_has_dez_component
        colleague_is_weekend_h8_or_dez_type = colleague_is_h8_type or colleague_has_dez_component

        if user_is_weekend_h8_or_dez_type and colleague_is_weekend_h8_or_dez_type:
            return True

    # --- Rule 4: "Op" matching (any day) ---
    if "Op" in user_parts and "Op" in colleague_parts:
        return True

    # --- Rule 5: T5 / T6 Equivalence Matching (any day) ---
    user_has_t5 = "T5" in user_parts
    user_has_t6 = "T6" in user_parts
    colleague_has_t5 = "T5" in colleague_parts
    colleague_has_t6 = "T6" in colleague_parts

    # If user's event has a T5 or T6 component AND
    # colleague's event also has a T5 or T6 component, they match.
    # This makes T5 equivalent to T6 for matching.
    if (user_has_t5 or user_has_t6) and \
            (colleague_has_t5 or colleague_has_t6):
        return True

    # --- Rule 6: Exact match for all other specific identities ---
    # This will cover cases like "J1" == "J1" or any other identical strings.
    if user_event_str == colleague_event_str:
        return True

    return False


def main():
    # 1) Create a hidden root window for Tk dialogs
    root = tk.Tk()
    root.withdraw()

    # 2) Ask user to select the Excel file
    excel_file = filedialog.askopenfilename(
        title="Odaberi Excel datoteku",
        filetypes=[("Excel datoteke", "*.xlsx *.xls"), ("Sve datoteke", "*.*")]
    )

    if not excel_file:
        print("Nije odabrana datoteka, pokušaj opet.")
        return

    excel_dir = os.path.dirname(excel_file)

    # 3) Load the selected Excel workbook
    try:
        wb = load_workbook(excel_file)
    except Exception as e:
        print(f"Pogreška pri učitavanju Excel datoteke: {e}")
        return

    sheet = wb.active

    # 4) Prompt user input for the surname
    valid_surname = False
    surname_row = None
    surname_input = ""  # Initialize surname_input

    while not valid_surname:
        surname_input = simpledialog.askstring(
            "Unos prezimena",
            "Unesi svoje prezime (kako si inače naveden u rasporedu):"
        )

        if surname_input is None:
            print("Nije uneseno prezime, probaj opet.")
            return

        for row_idx in range(1, sheet.max_row + 1):
            cell_value = sheet.cell(row=row_idx, column=1).value
            if cell_value and str(
                    cell_value).strip().lower() == surname_input.strip().lower():  # Case-insensitive match
                surname_row = row_idx
                # Use the exact casing from the sheet for the filename later, if desired, or stick to input
                # surname_input = str(cell_value).strip() # Optional: use exact name from sheet
                break

        if surname_row is None:
            messagebox.showerror(
                "Neispravan unos",
                f"Prezime '{surname_input}' nije pronađeno u prvom stupcu. Probaj opet."
            )
        else:
            valid_surname = True

    # 5) Determine the month name based on the first date in row=1, column=2 (Cell B1)
    first_date_cell = sheet.cell(row=1, column=2).value
    month = 0
    year = 0

    if isinstance(first_date_cell, str):
        try:
            day_str, month_str, year_str = first_date_cell.split(".")
            # day = int(day_str.strip()) # Day not strictly needed for month name
            month = int(month_str.strip())
            year = int(year_str.strip())
        except Exception:
            print("Ne mogu parsirati prvi datum (ćelija B1). Provjeri format npr. '1.2.2025'.")
            return
    elif isinstance(first_date_cell, (datetime.datetime, datetime.date)):  # Handle both datetime and date objects
        # day = first_date_cell.day
        month = first_date_cell.month
        year = first_date_cell.year
    else:
        print(f"Nepriznat format datuma u ćeliji B1: {first_date_cell} (tip: {type(first_date_cell)}).")
        return

    croatian_month_name = CRO_MONTHS.get(month, "Nepoznat_mjesec")

    # 6) Create ICS text
    ics_lines = []
    ics_lines.append("BEGIN:VCALENDAR")
    ics_lines.append("VERSION:2.0")
    ics_lines.append(f"PRODID:-//{surname_input}//Schedule Export//EN")  # Use user's name in PRODID

    max_col = sheet.max_column

    for col in range(2, max_col + 1):  # Start from column B (index 2)
        date_cell_value = sheet.cell(row=1, column=col).value
        if not date_cell_value:
            continue

        date_obj = None
        if isinstance(date_cell_value, str):
            try:
                d_str, m_str, y_str = date_cell_value.split(".")
                day_d = int(d_str.strip())
                month_d = int(m_str.strip())
                year_d = int(y_str.strip())
                date_obj = datetime.date(year_d, month_d, day_d)
            except Exception:
                # print(f"Skipping invalid date string in cell (1, {col}): {date_cell_value}")
                continue
        elif isinstance(date_cell_value, (datetime.datetime, datetime.date)):  # Handle both datetime and date
            date_obj = date_cell_value.date() if isinstance(date_cell_value, datetime.datetime) else date_cell_value
        else:
            # print(f"Skipping unrecognized date format in cell (1, {col}): {date_cell_value}")
            continue

        if date_obj is None:  # Should be caught above, but as a safeguard
            continue

        schedule_value = sheet.cell(row=surname_row, column=col).value
        if not schedule_value:
            continue

        user_schedule_str = str(schedule_value).strip()
        if not user_schedule_str:  # Skip if schedule is empty string after stripping
            continue

        uid_str = f"{surname_input}-{date_obj.year}-{date_obj.month}-{date_obj.day}-{col}-{uuid.uuid4()}@MyOrgSchedule"

        event_summary = user_schedule_str
        event_description_parts = []

        # Find colleagues if the event is not in the excluded list
        if user_schedule_str not in EXCLUDED_EVENTS_FOR_MATCHING:
            colleagues_found = []
            for r_idx in range(1, sheet.max_row + 1):
                if r_idx == surname_row:  # Don't match with self
                    continue

                colleague_name_cell = sheet.cell(row=r_idx, column=1).value
                if not colleague_name_cell:
                    continue
                colleague_name = str(colleague_name_cell).strip()

                colleague_schedule_cell = sheet.cell(row=r_idx, column=col).value
                if not colleague_schedule_cell:
                    continue

                colleague_schedule_str = str(colleague_schedule_cell).strip()
                if not colleague_schedule_str:
                    continue

                if is_match(user_schedule_str, colleague_schedule_str, date_obj):
                    colleagues_found.append(colleague_name)

            if colleagues_found:
                event_description_parts.append(f" {', '.join(colleagues_found)}")

        # --- Define event start and end times based on schedule_str ---
        ics_lines.append("BEGIN:VEVENT")
        ics_lines.append(f"UID:{uid_str}")
        ics_lines.append(
            f"DTSTAMP:{datetime.datetime.now(datetime.timezone.utc).strftime('%Y%m%dT%H%M%SZ')}")  # Use UTC for DTSTAMP

        if user_schedule_str in ["SD", "PD", "GO", "bol", "sist", "postdipl", "klaićeva"]:  # All-day events
            dt_start = date_obj.strftime("%Y%m%d")
            dt_end_obj = date_obj + datetime.timedelta(days=1)
            dt_end = dt_end_obj.strftime("%Y%m%d")
            ics_lines.append(f"DTSTART;VALUE=DATE:{dt_start}")
            ics_lines.append(f"DTEND;VALUE=DATE:{dt_end}")
        elif user_schedule_str.endswith("p"):  # 14:00 to 20:00
            start_dt = datetime.datetime(year=date_obj.year, month=date_obj.month, day=date_obj.day, hour=14, minute=0)
            end_dt = datetime.datetime(year=date_obj.year, month=date_obj.month, day=date_obj.day, hour=20, minute=0)
            ics_lines.append(f"DTSTART:{start_dt.strftime('%Y%m%dT%H%M%S')}")
            ics_lines.append(f"DTEND:{end_dt.strftime('%Y%m%dT%H%M%S')}")
            ics_lines.append(f"LOCATION:{LOCATION}")
        elif user_schedule_str == "H16":  # 15:30 to 07:30 next day
            start_dt = datetime.datetime(year=date_obj.year, month=date_obj.month, day=date_obj.day, hour=15, minute=30)
            next_day_obj = date_obj + datetime.timedelta(days=1)
            end_dt = datetime.datetime(year=next_day_obj.year, month=next_day_obj.month, day=next_day_obj.day, hour=7,
                                       minute=30)
            ics_lines.append(f"DTSTART:{start_dt.strftime('%Y%m%dT%H%M%S')}")
            ics_lines.append(f"DTEND:{end_dt.strftime('%Y%m%dT%H%M%S')}")
            ics_lines.append(f"LOCATION:{LOCATION}")
        else:  # Default: 07:30 to 15:00
            start_dt = datetime.datetime(year=date_obj.year, month=date_obj.month, day=date_obj.day, hour=7, minute=30)
            end_dt = datetime.datetime(year=date_obj.year, month=date_obj.month, day=date_obj.day, hour=15, minute=0)
            ics_lines.append(f"DTSTART:{start_dt.strftime('%Y%m%dT%H%M%S')}")
            ics_lines.append(f"DTEND:{end_dt.strftime('%Y%m%dT%H%M%S')}")
            ics_lines.append(f"LOCATION:{LOCATION}")

        ics_lines.append(f"SUMMARY:{event_summary}")
        if event_description_parts:
            # Join parts with a newline character for the ICS DESCRIPTION field
            full_description = "\n".join(event_description_parts)
            # Properly escape characters for ICS: backslashes, commas, semicolons, newlines
            full_description = full_description.replace("\\", "\\\\").replace(",", "\\,").replace(";", "\\;").replace(
                "\n", "\\n")
            ics_lines.append(f"DESCRIPTION:{full_description}")

        ics_lines.append("END:VEVENT")

    ics_lines.append("END:VCALENDAR")

    # 7) Write the ICS file
    ics_filename = f"{surname_input.replace(' ', '_')}_{croatian_month_name}_{year}.ics"  # Added year for uniqueness
    ics_full_path = os.path.join(excel_dir, ics_filename)

    try:
        with open(ics_full_path, "w", encoding="utf-8") as f:
            for line in ics_lines:
                f.write(line + "\r\n")  # Standard ICS line ending
        print(f"ICS datoteka '{ics_filename}' uspješno stvorena u direktoriju: {excel_dir}")
        messagebox.showinfo("Uspjeh", f"ICS datoteka '{ics_filename}' je uspješno stvorena!")
    except Exception as e:
        print(f"Pogreška pri pisanju ICS datoteke: {e}")
        messagebox.showerror("Pogreška", f"Pogreška pri pisanju ICS datoteke: {e}")
        return

    # 8) Open Google Calendar import page (changed from export to import)

    url_import = "https://calendar.google.com/calendar/r/settings/export"

    chrome_path_win_64 = r"C:/Program Files/Google/Chrome/Application/chrome.exe %s"
    chrome_path_win_32 = r"C:/Program Files (x86)/Google/Chrome/Application/chrome.exe %s"
    # For macOS, it might be:
    # chrome_path_mac = r"open -a /Applications/Google\ Chrome.app %s"

    opened_in_chrome = False
    try:
        webbrowser.get(chrome_path_win_64).open(url_import)  # Changed to url_import
        opened_in_chrome = True
    except webbrowser.Error:
        try:
            webbrowser.get(chrome_path_win_32).open(url_import)  # Changed to url_import
            opened_in_chrome = True
        except webbrowser.Error:
            pass  # Could not find Chrome at standard Windows paths

    if not opened_in_chrome:
        # Fallback for other OS or if Chrome not found by specific path
        # For macOS, you might try:
        # if sys.platform == "darwin": # import sys
        #     try:
        #         chrome_cmd = 'open -a "Google Chrome" %s'
        #         webbrowser.register('chrome', None, webbrowser.BackgroundBrowser(chrome_cmd))
        #         webbrowser.get('chrome').open(url_import)
        #         opened_in_chrome = True
        #     except webbrowser.Error:
        #         pass # Chrome not found or error opening

        # If still not opened in Chrome, use default browser
        if not opened_in_chrome:
            webbrowser.open(url_import)  # Changed to url_import
            print("Nije moguće pronaći Chrome na zadanom putu, otvaram u zadanom web-pregledniku...")
            messagebox.showinfo("Web-preglednik", "Otvaram Google Kalendar postavke u zadanom web-pregledniku za uvoz.")


if __name__ == "__main__":
    main()
